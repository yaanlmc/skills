#!/usr/bin/env python3
"""
工资表-人员核对脚本
用法: python check_headcount.py <Excel文件> [--sheets "表1;表2;..."]
"""
import sys
import argparse
import openpyxl

EXCLUDE = {
    '分表实发合计', '总表应发合计', '分表应发合计', '总表实发合计',
    '分表个税合计', '实发差值', '各项小计', '序号', '姓名',
    'None', '', 'none', 'N/A',
}


def find_name_col(ws):
    """Find the column containing '姓名' in sheet header rows."""
    for c in range(1, min(12, ws.max_column + 1)):
        h = str(ws.cell(1, c).value or '') + str(ws.cell(2, c).value or '')
        if '\u59d3\u540d' in h:
            return c
    # Fallback: scan first 3 data rows for non-empty string values
    for c in range(1, min(12, ws.max_column + 1)):
        v = ws.cell(3, c).value
        if v and '\u5c0f\u8ba1' not in str(v) and '\u5408\u8ba1' not in str(v):
            return c
    return 1


def find_dept_cols(ws):
    """Find 学部 (sector) and 部门 (dept) columns in the main sheet."""
    sector_col = None
    dept_col = None
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value or ''
        if '\u5b66\u90e8' in str(v) and '\u672a\u4f7f\u7528' not in str(v):
            sector_col = c
        if '\u90e8\u95e8' in str(v) and '\u771f\u5b9e' not in str(v) and '\u90e8\u95e8' in str(v):
            dept_col = c
    return sector_col, dept_col


def extract_names(ws, name_col, start_row=4):
    """Extract names from a sheet, excluding header/total rows."""
    names = {}
    for r in range(start_row, ws.max_row + 1):
        v = ws.cell(r, name_col).value
        if v and str(v).strip() and str(v).strip() not in EXCLUDE:
            names[str(v).strip()] = r
    return names


def main():
    parser = argparse.ArgumentParser(description='\u5de5\u8d44\u8868\u4eba\u5458\u2661\u6838\u5bf9')
    parser.add_argument('file', help='Excel\u6587\u4ef6\u8def\u5f84')
    parser.add_argument('--sheets', default='',
                        help='\u8981\u6838\u5bf9\u7684\u5206\u8868\u540d\uff08\u7528\u5206\u53f7\u5206\u9694\uff0c\u5982"\u3010\u98df\u5802\u3011-\u3010\u5c0f\u5b66\u3011;\u3010\u603b\u52a1\u3011-\u3010\u5c0f\u5b66\u3011"\uff0c\u9ed8\u8ba4\u81ea\u52a8\u8bc6\u522b\u540e\u7aef\u5206\u8868"')
    parser.add_argument('--main-sheet', default='\u6c47\u603b\u8868',
                        help='\u4e3b\u8868\u540d\u79f0\uff0c\u9ed8\u8ba4\u6c47\u603b\u8868')
    args = parser.parse_args()

    sys.stdout.reconfigure(encoding='utf-8')

    wb = openpyxl.load_workbook(args.file, data_only=True)
    print(f'=== \u6587\u4ef6: {args.file} ===')
    print(f'\u5de5\u4f5c\u8868: {wb.sheetnames}')

    # Identify main sheet
    main_ws = None
    for name in wb.sheetnames:
        if args.main_sheet in name or name == args.main_sheet:
            main_ws = wb[name]
            break
    if main_ws is None:
        main_ws = wb[wb.sheetnames[0]]

    main_name_col = find_name_col(main_ws)
    main_names = extract_names(main_ws, main_name_col, start_row=4)
    sector_col, dept_col = find_dept_cols(main_ws)
    print(f'\n[\u6c47\u603b\u8868] {main_ws.title} \u4eba\u6570: {len(main_names)}')

    # Identify sub sheets
    if args.sheets:
        sub_sheets = [s.strip() for s in args.sheets.split(';') if s.strip()]
    else:
        # Default: last 6 sheets
        sub_sheets = wb.sheetnames[-6:] if len(wb.sheetnames) >= 6 else wb.sheetnames[1:]

    print(f'[\u5206\u8868] {sub_sheets}')

    # Collect names from all sub sheets
    all_sub_names = {}
    sheet_breakdown = {}
    for sname in sub_sheets:
        if sname not in wb.sheetnames:
            print(f'  [!] \u8868\u4e0d\u5b58\u5728: {sname}')
            continue
        ws = wb[sname]
        nc = find_name_col(ws)
        names = extract_names(ws, nc, start_row=1)
        sheet_breakdown[sname] = names
        all_sub_names.update(names)
        print(f'  {sname}: {len(names)} \u4eba')

    # Statistics
    total_sub = len(all_sub_names)
    print(f'\n\u6c47\u603b\u8868: {len(main_names)} \u4eba  |  \u5206\u8868\u5408\u8ba1(\u53bb\u91cd): {total_sub} \u4eba')

    # Difference analysis
    only_main = {n: v for n, v in main_names.items() if n not in all_sub_names}
    only_sub = {n: v for n, v in all_sub_names.items() if n not in main_names}

    print(f'\n--- \u53ea\u5728\u6c47\u603b\u8868 (\u4e0d\u5728\u5206\u8868): {len(only_main)} \u4eba ---')
    if only_main:
        # Show dept info for each orphan
        for name in sorted(only_main.keys()):
            row = main_names[name]
            sect = main_ws.cell(row, sector_col).value if sector_col else ''
            dept = main_ws.cell(row, dept_col).value if dept_col else ''
            bank = main_ws.cell(row, 48).value or ''
            print(f'  {name} | \u5b66\u90e8:{sect} | \u90e8\u95e8:{dept} | \u94f6\u884c\u5361:{bank}')

    print(f'\n--- \u53ea\u5728\u5206\u8868 (\u4e0d\u5728\u6c47\u603b\u8868): {len(only_sub)} \u4eba ---')
    if only_sub:
        for name in sorted(only_sub.keys()):
            sheets = [s for s, ns in sheet_breakdown.items() if name in ns]
            print(f'  {name} | \u8868: {sheets}')

    if not only_main and not only_sub:
        print('\u2705 \u5b8c\u5168\u4e00\u81f4')


if __name__ == '__main__':
    main()
